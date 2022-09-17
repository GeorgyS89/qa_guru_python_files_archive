import codecs
import csv
import os
import zipfile

from PyPDF2 import PdfReader
from openpyxl import load_workbook

from archive import create_archive, move_archive

create_archive()
move_archive()

def test_archived_csv():
    with zipfile.ZipFile(os.path.abspath('./resources/archive.zip')) as test_archive:
        with test_archive.open('resources/Employees.csv') as test_csv:
            table = csv.reader(codecs.iterdecode(test_csv, 'utf-8'))
            for line_number, line in enumerate(table, 1):
                if line_number == 3:
                    assert line[3] == '123458'


def test_archived_xlsx():
    with zipfile.ZipFile(os.path.abspath('./resources/archive.zip')) as test_archive:
        with test_archive.open('resources/Grades.xlsx') as test_xlsx:
            test_xlsx = load_workbook(test_xlsx)
            sheet = test_xlsx.active
            data = sheet.cell(row=8, column=1).value
            assert data == 'Журавлев Антон Владимирович'

def test_archived_pdf():
    with zipfile.ZipFile(os.path.abspath('./resources/archive.zip')) as test_archive:
        with test_archive.open('resources/ISTQB.pdf') as test_pdf:
            test_pdf = PdfReader(test_pdf)
            num_pages = len(test_pdf.pages)
            assert num_pages == 96



